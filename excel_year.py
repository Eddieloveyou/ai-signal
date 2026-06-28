#!/usr/bin/env python3
# Per-year-sheet workbook for the hard-gated variant.
import os, pickle, numpy as np, pandas as pd
import engine as E
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

HERE=os.path.dirname(os.path.abspath(__file__))
RES=pickle.load(open(os.path.join(HERE,'res_gated.pkl'),'rb'))
FEAT=E.FEAT; C2N=E.C2N
days=RES['days']; dec=RES['decisions']; nv=RES['nv']; dret=RES['daily_ret']
REQUIRE_BULL=True
def fmtd(d): return f"{d[:4]}-{d[4:6]}-{d[6:]}"

def top5_for(d):
    df=d['df']
    if df is None or len(df)==0: return None
    sub=df[(df.tiegao)&(df.amount_yi>=1.0)&(df.kpzf<0.098)&(~df.st)&df.kpzf.notna()].copy()
    return sub.sort_values('liangbi',ascending=False).head(5)

def reason(row,d):
    if row.code in d['picks']: return ''
    if d['regime']=='防守': return '当日防守空仓'
    if not (pd.notna(row.liangbi) and 3<row.liangbi<20):
        if d['engine'] in ('兜底','深兜底'): return f'{d["engine"]}日·量比未居首'
        return '量比不在(3,20)'
    if getattr(row,'hi_yin',False) and pd.notna(getattr(row,'hi_ds',np.nan)) and row.hi_ds>7:
        return '阴线高点·距高点>7天剔除'
    if REQUIRE_BULL and not row.bull: return '非多头排列'
    if not (pd.notna(row.kpzf) and row.kpzf<0.098): return '开盘涨幅≥9.8%'
    if pd.notna(row.chg20) and row.chg20>1.0: return '20日涨幅>100%剔除'
    if pd.notna(row.chg20) and row.chg20>0.8: return '20日涨幅>80%(主选二三剔除)'
    if pd.notna(row.kpzf) and row.kpzf<=-0.015: return '开盘涨幅≤-1.5%(主选二三剔除)'
    return '量比未进前三/份额已满'

HEAD=['排名','代码','名称','量比','贴高比','20日涨幅','竞价最终价','竞价区间高','竞价区间低','早盘指示价','终值位置','开盘涨幅','成交额(亿)','T+1卖出收益','未买入原因']
NCOL=len(HEAD)
hfill=PatternFill('solid',fgColor='1F4E78'); hfont=Font(color='FFFFFF',bold=True,size=10)
attackfill=PatternFill('solid',fgColor='FCE4D6'); defendfill=PatternFill('solid',fgColor='E2EFDA')
infofill=PatternFill('solid',fgColor='FFF2CC'); star=Font(bold=True)  # bought row: bold only, no color
thin=Side(style='thin',color='D9D9D9'); border=Border(left=thin,right=thin,top=thin,bottom=thin)
center=Alignment(horizontal='center',vertical='center'); left=Alignment(horizontal='left',vertical='center')
widths=[10,11,10,8,9,10,11,11,11,11,9,9,10,11,30]

def realized_ret(k_global, d):
    # actual return of the day's picks: buy T open (qfq), sell T+1 close (qfq)
    if not d['picks'] or k_global+1>=len(days): return None
    T=days[k_global]; T1=days[k_global+1]; rr=[]
    for c in d['picks']:
        f=FEAT[c]
        if T in f.index and T1 in f.index and f.loc[T,'oq']>0:
            sellp=f.loc[T1,'oq'] if (E.limit_down_T(c,T) and f.loc[T1,'oq']>0) else f.loc[T1,'cq']
            rr.append(sellp/f.loc[T,'oq']-1)
    return float(np.mean(rr)) if rr else None

def build_year_sheet(ws, decs, ysum):
    # ysum: (year, yret, env, maxdd, ntr)
    y,yret,env,maxdd,ntr=ysum
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=NCOL)
    s=ws.cell(1,1,f"{y}年  ·  年收益率 {yret*100:+.1f}%  ·  期末净值 {env:.2f}  ·  年内最大回撤 {maxdd*100:.1f}%  ·  交易 {ntr} 笔")
    s.fill=infofill; s.font=Font(bold=True,size=12,color='C00000' if yret>0 else '008000'); s.alignment=center
    for j in range(1,NCOL+1): ws.cell(1,j).border=border
    for j,h in enumerate(HEAD,1):
        c=ws.cell(2,j,h); c.fill=hfill; c.font=hfont; c.alignment=center; c.border=border
    ws.freeze_panes='A3'; r=3
    for k_global,d in decs:
        dd=d['date']; dr=dret.loc[dd]; cur=nv.loc[dd]
        picks_names='、'.join(C2N[c] for c in d['picks']) if d['picks'] else '—'
        npick=len(d['picks']); batch='批A' if k_global%2==0 else '批B'
        pos=('空仓' if d['regime']=='防守' else (f'50%×{batch} · {npick}只均分' if npick else f'{batch}·无合格标的(空仓)'))
        rret=realized_ret(k_global,d)
        rret_txt=(f"本笔买入收益:{rret*100:+.2f}%" if rret is not None else "本笔买入收益:—")
        info=(f"{fmtd(dd)}  |  {d['regime']}  |  引擎:{d['engine']}  |  仓位:{pos}  |  "
              f"{rret_txt}  |  实际买入:{picks_names}")
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=NCOL)
        ic=ws.cell(r,1,info); ic.fill=attackfill if d['regime']=='进攻' else defendfill
        # font color driven by ACTUAL buy return (本笔买入收益): negative -> green, else plain black (no red)
        ic.font=Font(bold=True,color=('008000' if (rret is not None and rret<0) else '000000'),size=10); ic.alignment=left
        for j in range(1,NCOL+1): ws.cell(r,j).border=border
        r+=1
        t5=top5_for(d)
        if t5 is not None and len(t5):
            for rank,(_,row) in enumerate(t5.iterrows(),1):
                bought=row.code in d['picks']; mark=f'{rank} ★' if bought else f'{rank}'
                tg=row.tm1_close/row.high60 if (pd.notna(row.high60) and row.high60>0) else None
                vals=[mark,row.code,row['name'],
                      round(row.liangbi,2) if pd.notna(row.liangbi) else None,
                      round(tg,4) if tg is not None else None,
                      round(row.chg20,4) if pd.notna(row.chg20) else None,
                      round(row.a_close,2) if pd.notna(row.a_close) else None,
                      round(row.a_high,2) if pd.notna(row.a_high) else None,
                      round(row.a_low,2) if pd.notna(row.a_low) else None,
                      round(row.a_open,2) if pd.notna(row.a_open) else None,
                      round(row.zwpos,4) if pd.notna(row.zwpos) else None,
                      round(row.kpzf,4) if pd.notna(row.kpzf) else None,
                      round(row.amount_yi,2) if pd.notna(row.amount_yi) else None,
                      None,  # placeholder for T+1 return, filled below
                      reason(row,d)]
                # T+1 sell return for this candidate (buy T open, sell T+1 close, qfq)
                rt1=None
                if k_global+1<len(days):
                    fc=FEAT[row.code]; T=days[k_global]; T1=days[k_global+1]
                    if T in fc.index and T1 in fc.index and fc.loc[T,'oq']>0:
                        rt1=fc.loc[T1,'cq']/fc.loc[T,'oq']-1
                vals[13]=round(rt1,4) if rt1 is not None else None
                for j,v in enumerate(vals,1):
                    c=ws.cell(r,j,v); c.border=border; c.alignment=left if j in (3,15) else center
                    if bought: c.font=star  # bold only, no color
                for cc in (4,7,8,9,10,13): ws.cell(r,cc).number_format='0.00'
                for cc in (5,6,11,12,14): ws.cell(r,cc).number_format='0.0%'
                r+=1
        else:
            c=ws.cell(r,1,'(当日无贴60日高的合格候选)'); c.alignment=left
            for j in range(1,NCOL+1): ws.cell(r,j).border=border
            r+=1
        r+=1
    for j,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(j)].width=w

def year_stats(y):
    didx=[i for i,dd in enumerate(days) if dd.startswith(y)]
    if not didx: return None
    snv=nv.iloc[didx[0]-1] if didx[0]>0 else 1.0; env=nv.iloc[didx[-1]]; yret=env/snv-1
    seg=nv.iloc[didx[0]:didx[-1]+1].values; arr=np.concatenate([[snv],seg])
    dd_=(arr/np.maximum.accumulate(arr)-1).min(); ntr=sum(1 for i in didx if dec[i]['picks'])
    return (y,yret,float(env),float(dd_),ntr)

wb=Workbook(); wb.remove(wb.active)
years=['2021','2022','2023','2024','2025','2026']
YSTATS={y:year_stats(y) for y in years}
for y in years:
    decs=[(k,d) for k,d in enumerate(dec) if d['date'].startswith(y)]
    if not decs or YSTATS[y] is None: continue
    ws=wb.create_sheet(y)
    build_year_sheet(ws, decs, YSTATS[y])

# 逐年收益率 sheet
ws2=wb.create_sheet('逐年收益率')
ws2.merge_cells('A1:E1'); t=ws2.cell(1,1,'AI主池打板 · 逐年收益率(硬过滤:高开>9.8%/20日涨幅>100%/量比>20 全档剔除)')
t.font=Font(bold=True,size=13); t.alignment=center
for j,h in enumerate(['年份','年收益率','期末净值','年内最大回撤','交易笔数'],1):
    c=ws2.cell(3,j,h); c.fill=hfill; c.font=hfont; c.alignment=center; c.border=border
rr=4
for y in years:
    didx=[i for i,dd in enumerate(days) if dd.startswith(y)]
    if not didx: continue
    snv=nv.iloc[didx[0]-1] if didx[0]>0 else 1.0; env=nv.iloc[didx[-1]]; yret=env/snv-1
    seg=nv.iloc[didx[0]:didx[-1]+1].values; arr=np.concatenate([[snv],seg])
    dd_=(arr/np.maximum.accumulate(arr)-1).min(); ntr=sum(1 for i in didx if dec[i]['picks'])
    for j,v in enumerate([y,yret,round(env,4),dd_,ntr],1):
        c=ws2.cell(rr,j,v); c.border=border; c.alignment=center
    ws2.cell(rr,2).number_format='0.0%'; ws2.cell(rr,2).font=Font(color='C00000' if yret>0 else '008000')
    ws2.cell(rr,4).number_format='0.0%'; ws2.cell(rr,4).font=Font(color='008000'); rr+=1
maxdd=(nv.values/np.maximum.accumulate(nv.values)-1).min()
for j,v in enumerate(['全程',nv.iloc[-1]-1,round(nv.iloc[-1],2),maxdd,sum(1 for d in dec if d['picks'])],1):
    c=ws2.cell(rr,j,v); c.fill=infofill; c.font=Font(bold=True); c.border=border; c.alignment=center
ws2.cell(rr,2).number_format='0.0%'; ws2.cell(rr,2).font=Font(bold=True,color='C00000'); ws2.cell(rr,4).number_format='0.0%'
notes=['说明:','① 每年一个工作表;硬过滤(高开>9.8%/20日涨幅>100%/量比>20)贯穿全部档位及Top5。',
       '② 股票池=当下AI主线60只,存在幸存者偏差,绝对收益偏高,不代表实盘可达。',
       '③ 未计佣金/印花税/竞价涨停买不进等摩擦。净值=两等份错开一天滚动之和(各0.5),T开盘买T+1收盘卖,防守空仓。',
       '④ 量比=(竞价量/100)/(过去5日日均量/240);成交额=T-1日成交额≥1亿元;保留多头排列过滤。',
       '⑤ 新增过滤:60日高点那根K线若为阴线,则要求距高点≤7个交易日,否则剔除(阳线高点不限距离)。',
       '   依据:阴线高点>7天=久攻不破走弱(每笔+0.94%/胜率48%);阳线高点远距离反而是长底盘整再起(优质)。',
       '⑥ 深兜底加量比>5下限:深兜底为最弱档(胜率<50%、几乎不贡献收益),只取量比>5,无则空仓→回撤最优(−25.7%)、收益几乎无损。']
for i,nt in enumerate(notes):
    ws2.cell(5+rr-4+i+1,1,nt).font=Font(size=10,italic=(i>0))
for j,w in enumerate([10,12,12,14,10],1): ws2.column_dimensions[get_column_letter(j)].width=w

# 净值与指数 data + charts
ws3=wb.create_sheet('净值与指数'); ws3.append(['日期','组合净值','等权指数','指数MA20'])
idxl=RES['idx_level']; idxm=RES['idx_ma20']
for dd in days:
    ws3.append([fmtd(dd),float(nv.loc[dd]),
                float(idxl.loc[dd]) if dd in idxl.index else None,
                float(idxm.loc[dd]) if dd in idxm.index and pd.notna(idxm.loc[dd]) else None])
n=len(days)
ch1=LineChart(); ch1.title='组合净值曲线'; ch1.height=8; ch1.width=18
ch1.add_data(Reference(ws3,min_col=2,min_row=1,max_row=n+1),titles_from_data=True)
ch1.set_categories(Reference(ws3,min_col=1,min_row=2,max_row=n+1)); ws2.add_chart(ch1,'G3')
ch2=LineChart(); ch2.title='等权指数 vs MA20'; ch2.height=8; ch2.width=18
ch2.add_data(Reference(ws3,min_col=3,max_col=4,min_row=1,max_row=n+1),titles_from_data=True)
ch2.set_categories(Reference(ws3,min_col=1,min_row=2,max_row=n+1)); ws2.add_chart(ch2,'G20')
ws3.column_dimensions['A'].width=12
for col in 'BCD': ws3.column_dimensions[col].width=12

OUT=os.path.join(HERE,'AI主池打板回溯_2021-2026_阴线距离过滤版.xlsx')
wb.save(OUT); print('saved',OUT,'sheets=',wb.sheetnames)
