#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""AI 主池打板回溯引擎 v2.1 (2021-01-01 ~ 最新交易日)

逐日重放 60 只 AI 主池, 复算同花顺口径量比 + 关键指标, 套用全局硬过滤与
主选/兜底/深兜底选股, 计算两份错开滚动净值与逐年收益率, 并输出统一美化的
Excel 工作簿(每年一个工作表)。

用法:
  export TUSHARE_TOKEN=你的token
  python3 backtest.py                  # 回溯到最新交易日, 输出 ai_backtest_2021_2026.xlsx
  python3 backtest.py 20210101 20261231 out.xlsx   # 自定义区间与文件名

数据口径(全程前复权, 一致):
  · 行情: Tushare daily(原始) × adj_factor → 前复权; 量比/成交额用原始量额。
  · 集合竞价: Tushare stk_auction_o(9:15-9:25 开高低收量), 竞价最终价 = open = 当日开盘价;
    竞价量(股)在 vol 字段。历史可回溯, 是本回测唯一可得的竞价数据源。
  · ST: namechange 历史名称区间, 名称含 'ST' 即视为当日 ST。
  · 择时: 60 只等权指数(日收益等权)对比其 MA20。
  · ETF 参考: 银行ETF 512800.SH / 黄金ETF 518880.SH(仅参考列, 不实际持有)。

引擎规则见 strategy_spec.txt(v2.1)。本回测含固定池幸存者偏差, 账面收益仅供复盘。
"""
import os, sys, time
import numpy as np, pandas as pd
from concurrent.futures import ThreadPoolExecutor

# ─────────────────────────── 配置 ───────────────────────────
LOW_OPEN_TOL = -0.015          # 选2/选3 允许的最大低开幅度(gap > 此值才保留)
RUN20_CAP_MAIN = 1.00          # 20日涨幅硬上限(>则全档剔除)
RUN20_CAP_SUB = 0.80           # 选2/选3 的 20日涨幅上限
AMT_MIN = 100000.0             # 成交额 ≥ 1 亿 (daily.amount 单位千元 → 1亿=100000)
QB_LO, QB_HI = 3.0, 18.0       # 候选量比区间
QB_HARD = 20.0                 # 硬过滤: 量比 > 20 剔除
QB_DEEP_LO = 5.0               # 深兜底: 量比 ∈ (5,18)
NEAR_HIGH = 0.93               # 贴高: T-1收盘 ≥ 60日最高 × 93%
GAP_CAP = 0.098                # 开盘涨幅硬上限 9.8%

POOL = {
 '风华高科':'000636.SZ','华工科技':'000988.SZ','德明利':'001309.SZ','大族激光':'002008.SZ','科大讯飞':'002230.SZ',
 '东山精密':'002384.SZ','双环传动':'002472.SZ','科士达':'002518.SZ','英维克':'002837.SZ','洁美科技':'002859.SZ',
 '深南电路':'002916.SZ','创世纪':'300083.SZ','拓尔思':'300229.SZ','中际旭创':'300308.SZ','润泽科技':'300442.SZ',
 '胜宏科技':'300476.SZ','高澜股份':'300499.SZ','新易盛':'300502.SZ','昊志机电':'300503.SZ','长芯博创':'300548.SZ',
 '精测电子':'300567.SZ','太辰光':'300570.SZ','长川科技':'300604.SZ','光库科技':'300620.SZ','江苏雷利':'300660.SZ',
 '江丰电子':'300666.SZ','金力永磁':'300748.SZ','龙磁科技':'300835.SZ','兆龙互连':'300913.SZ','申菱环境':'301018.SZ',
 '铜冠铜箔':'301217.SZ','永鼎股份':'600105.SH','中国巨石':'600176.SH','生益科技':'600183.SH','有研新材':'600206.SH',
 '亨通光电':'600487.SH','中天科技':'600522.SH','柏诚股份':'601133.SH','长飞光纤':'601869.SH','中科曙光':'603019.SH',
 '华正新材':'603186.SH','景旺电子':'603228.SH','宏和科技':'603256.SH','鸣志电器':'603728.SH','兆易创新':'603986.SH',
 '中微公司':'688012.SH','绿的谐波':'688017.SH','拓荆科技':'688072.SH','步科股份':'688160.SH','生益电子':'688183.SH',
 '寒武纪':'688256.SH','联瑞新材':'688300.SH','汇成股份':'688403.SH','源杰科技':'688498.SH','佰维存储':'688525.SH',
 '华丰科技':'688629.SH','鼎通科技':'688668.SH','伟创电气':'688698.SH','普冉股份':'688766.SH','中控技术':'688777.SH'}
CODE2NM = {v: k for k, v in POOL.items()}

# 硬过滤特例名单(按代码)
HIGH_OPEN_BAN = {'002837.SZ', '603986.SH', '688300.SH', '300499.SZ', '688017.SH'}  # 高开不买(gap≥0剔)
LOW_OPEN_BAN  = {'688183.SH', '000636.SZ', '603228.SH', '600487.SH'}              # 低开不买(gap<0剔)
ETF = {'银行ETF': '512800.SH', '黄金ETF': '518880.SH'}

pro = None      # tushare pro_api 句柄, 在 init_pro() 中惰性初始化

def init_pro():
    global pro
    if pro is not None:
        return
    import tushare as ts
    if not os.environ.get('TUSHARE_TOKEN'):
        sys.exit('缺少 TUSHARE_TOKEN 环境变量')
    ts.set_token(os.environ['TUSHARE_TOKEN'])
    pro = ts.pro_api()

def is_chuangchuang(code):
    """双创(创业板300/科创板688) 跌停阈值-19.8%; 其余主板-9.8%。"""
    return code[:3] in ('300', '688')

# ─────────────────────────── 取数 ───────────────────────────
def _retry(fn, *a, **k):
    for i in range(4):
        try:
            r = fn(*a, **k)
            if r is not None:
                return r
        except Exception:
            time.sleep(0.6 * (i + 1))
    return None

def fetch_daily(code, start, end):
    """原始日线 + adj_factor → 前复权(以区间内最新因子为基准, 全程一致)。"""
    d = _retry(pro.daily, ts_code=code, start_date=start, end_date=end)
    if d is None or not len(d):
        return code, None
    af = _retry(pro.adj_factor, ts_code=code, start_date=start, end_date=end)
    d = d.sort_values('trade_date').reset_index(drop=True)
    if af is not None and len(af):
        af = af[['trade_date', 'adj_factor']].sort_values('trade_date')
        d = d.merge(af, on='trade_date', how='left')
        d['adj_factor'] = d['adj_factor'].ffill().bfill()
        base = d['adj_factor'].iloc[-1]                    # 前复权基准 = 最新因子
        ratio = d['adj_factor'] / base
        for col in ('open', 'high', 'low', 'close', 'pre_close'):
            d[f'q{col}'] = d[col] * ratio
    else:                                                  # 无因子则视同不复权
        for col in ('open', 'high', 'low', 'close', 'pre_close'):
            d[f'q{col}'] = d[col]
    return code, d

def fetch_auction(code, start, end):
    """stk_auction_o 集合竞价(分年取, 规避单次行数上限)。"""
    frames = []
    y0, y1 = int(start[:4]), int(end[:4])
    for y in range(y0, y1 + 1):
        s = max(start, f'{y}0101'); e = min(end, f'{y}1231')
        x = _retry(pro.stk_auction_o, ts_code=code, start_date=s, end_date=e)
        if x is not None and len(x):
            frames.append(x)
    if not frames:
        return code, None
    a = pd.concat(frames).drop_duplicates('trade_date').sort_values('trade_date')
    return code, a.set_index('trade_date')

def fetch_st_intervals(code):
    """返回 [(start_date, end_date或None), ...] 为 ST 名称的区间。"""
    nc = _retry(pro.namechange, ts_code=code)
    if nc is None or not len(nc):
        return code, []
    out = []
    for _, r in nc.iterrows():
        if 'ST' in str(r.get('name', '')):
            out.append((str(r['start_date']), str(r['end_date']) if r.get('end_date') else None))
    return code, out

def fetch_etf(code, start, end):
    d = _retry(pro.fund_daily, ts_code=code, start_date=start, end_date=end)
    if d is None or not len(d):
        return code, None
    return code, d.sort_values('trade_date').set_index('trade_date')

def trade_days(start, end):
    cal = _retry(pro.trade_cal, exchange='SSE', start_date=start, end_date=end, is_open='1')
    return sorted(cal['cal_date'].tolist())

def is_st_on(intervals, date):
    for s, e in intervals:
        if s <= date and (e is None or date <= e):
            return True
    return False

# ─────────────────────────── 引擎 ───────────────────────────
def build():
    init_pro()
    end = sys.argv[2] if len(sys.argv) > 2 and sys.argv[2][:8].isdigit() else time.strftime('%Y%m%d')
    start_bt = sys.argv[1] if len(sys.argv) > 1 and sys.argv[1][:8].isdigit() else '20210101'
    fetch_start = '20200801'        # 多取半年做 MA60 / 60日高 / 量比的回看窗口
    print(f'回溯区间 {start_bt} ~ {end}; 拉取起点 {fetch_start}')

    codes = list(POOL.values())
    daily, auction, st_iv = {}, {}, {}
    with ThreadPoolExecutor(max_workers=12) as ex:
        for c, d in ex.map(lambda c: fetch_daily(c, fetch_start, end), codes):
            if d is not None: daily[c] = d
        for c, a in ex.map(lambda c: fetch_auction(c, fetch_start, end), codes):
            if a is not None: auction[c] = a
        for c, iv in ex.map(fetch_st_intervals, codes):
            st_iv[c] = iv
    etf = {}
    for nm, c in ETF.items():
        _, d = fetch_etf(c, start_bt, end)
        etf[nm] = d
    print(f'取数完成: daily {len(daily)} 只, auction {len(auction)} 只')

    # 全部交易日(以日线并集为准)
    all_days = sorted(set().union(*[set(d['trade_date']) for d in daily.values()]))
    days = [t for t in all_days if start_bt <= t <= end]

    # ── 等权指数(日收益等权)+ MA20 → 择时 ──
    qclose = pd.DataFrame({c: d.set_index('trade_date')['qclose'] for c, d in daily.items()})
    qclose = qclose.sort_index()
    rets = qclose.pct_change()
    idx_ret = rets.mean(axis=1)                 # 当日全池等权日收益
    idx_nav = (1 + idx_ret.fillna(0)).cumprod()
    idx_ma20 = idx_nav.rolling(20).mean()

    # 预建每只票按日期索引的视图, 便于 T / T-1 / T+1 取值
    dview = {c: d.set_index('trade_date') for c, d in daily.items()}

    pos_of = {d: i for i, d in enumerate(all_days)}

    records = []                                 # 每个交易日一条
    for T in days:
        i = pos_of[T]
        prev = all_days[i - 1] if i > 0 else None
        nxt = all_days[i + 1] if i + 1 < len(all_days) else None
        if prev is None:
            continue
        regime = 'attack' if (T in idx_nav.index and prev in idx_ma20.index
                              and not np.isnan(idx_ma20.get(prev, np.nan))
                              and idx_nav.get(prev) > idx_ma20.get(prev)) else 'defend'
        rec = dict(T=T, prev=prev, nxt=nxt, regime=regime, rows=[], picks=[], tier='', trade_ret=None)
        if regime == 'defend':
            records.append(rec); continue

        # 计算全池当日候选指标
        rows = []
        for c, dv in dview.items():
            if prev not in dv.index or T not in dv.index:
                continue
            hist = dview[c].loc[:prev]                       # 截至 T-1 的历史(含 T-1)
            if len(hist) < 61:
                continue
            au = auction.get(c)
            if au is None or T not in au.index:
                continue
            av = au.loc[T, 'vol']                            # 竞价量(股)
            qcl = hist['qclose'].values                      # 前复权收盘序列(含 T-1)
            vo = hist['vol'].values                          # 原始量(手)
            ctm1 = qcl[-1]
            v5 = vo[-5:].mean()
            if pd.isna(av) or not av or v5 <= 0 or len(qcl) < 21:
                continue
            qb = (av / 100) / (v5 / 240)                     # 同花顺口径量比
            permin5 = v5 / 240                               # 近5日每分钟均量(手)

            # 竞价最终价 = 当日开盘价(用 T 的前复权开盘, 与 ctm1 同空间)
            px = dv.loc[T, 'qopen']
            if pd.isna(px) or not ctm1:
                continue
            gap = px / ctm1 - 1
            # 终值位置(仅展示): (open - low)/(high - low); 高=低记 50%
            ahi, alo = au.loc[T, 'high'], au.loc[T, 'low']
            aop = au.loc[T, 'open']
            pos = 0.5 if (ahi is None or alo is None or ahi == alo) else (aop - alo) / (ahi - alo)

            h60 = hist['qhigh'].values[-60:].max()
            ma20 = qcl[-20:].mean(); ma60 = qcl[-60:].mean()
            run20 = qcl[-1] / qcl[-21] - 1
            amt = hist['amount'].values[-1]                  # T-1 成交额(千元)
            near = ctm1 >= NEAR_HIGH * h60
            trend = (ctm1 > ma20) and (ma20 > ma60)

            # 硬过滤4: 60日最高K线为阴线 且 T-1距该高点 > 7个交易日
            hi60_arr = hist['qhigh'].values[-60:]
            jmax = int(np.argmax(hi60_arr))                  # 60日窗口内最高所在位置
            bar_open = hist['qopen'].values[-60:][jmax]
            bar_close = hist['qclose'].values[-60:][jmax]
            dist = (len(hi60_arr) - 1) - jmax                # T-1 距该高点的交易日数
            yin_high_bad = (bar_close < bar_open) and (dist > 7)

            isST = is_st_on(st_iv.get(c, []), T)

            # 全局硬过滤(任一命中即整票剔除)
            hard = []
            if gap > GAP_CAP: hard.append('开盘>9.8%')
            if run20 > RUN20_CAP_MAIN: hard.append('20涨>100%')
            if qb > QB_HARD: hard.append('量比>20')
            if yin_high_bad: hard.append('阴线高点>7日')
            if c in HIGH_OPEN_BAN and gap >= 0: hard.append('高开不买')
            if c in LOW_OPEN_BAN and gap < 0: hard.append('低开不买')
            if isST: hard.append('ST')

            rows.append(dict(c=c, nm=CODE2NM[c], qb=qb, permin5=permin5, gap=gap, pos=pos,
                             run20=run20, amt=amt, near=near, trend=trend, isST=isST,
                             hard=hard, hard_ok=(len(hard) == 0)))
        rec['rows'] = rows
        records.append(rec)

    # ── 选股(主选/兜底/深兜底)+ 逐笔实现收益(用于排名与净值) ──
    def real_ret(c, T, nxt):
        """该票在 T 的买入到卖出实现收益(前复权): T开盘买; 正常T+1收盘卖;
           若 T 收盘跌停则改 T+1 开盘卖。"""
        dv = dview.get(c)
        if dv is None or T not in dv.index or nxt is None or nxt not in dv.index:
            return None
        buy = dv.loc[T, 'qopen']
        # T 是否跌停: 原始 pct = close/pre_close - 1
        cl_t = dv.loc[T, 'close']; pc_t = dv.loc[T, 'pre_close']
        pct_t = (cl_t / pc_t - 1) if pc_t else 0
        limit = -0.198 if is_chuangchuang(c) else -0.098
        sell = dv.loc[nxt, 'qopen'] if pct_t <= limit else dv.loc[nxt, 'qclose']
        if not buy:
            return None
        return sell / buy - 1

    for rec in records:
        if rec['regime'] != 'attack' or not rec['rows']:
            continue
        T, nxt = rec['T'], rec['nxt']
        cand = [r for r in rec['rows'] if r['hard_ok'] and r['near'] and r['trend']
                and r['amt'] >= AMT_MIN and (QB_LO < r['qb'] < QB_HI) and (r['gap'] < GAP_CAP)
                and not r['isST']]
        cand.sort(key=lambda r: r['qb'], reverse=True)

        picks, tier = [], ''
        if cand:
            tier = '主选'
            picks.append(('选1', cand[0]))                          # 量比第1, 低开照买
            if len(cand) >= 2:
                r2 = cand[1]
                if r2['run20'] <= RUN20_CAP_SUB and r2['gap'] > LOW_OPEN_TOL:
                    picks.append(('选2', r2))
            if len(cand) >= 3:
                r3 = cand[2]
                if r3['run20'] <= RUN20_CAP_SUB and r3['gap'] > LOW_OPEN_TOL:
                    picks.append(('选3', r3))
        else:
            # 兜底: 去量比>3下限与20涨≤80限制; 留贴高+多头+额≥1亿+非ST(仍受硬过滤)
            fb = [r for r in rec['rows'] if r['hard_ok'] and r['near'] and r['trend']
                  and r['amt'] >= AMT_MIN and not r['isST']]
            fb.sort(key=lambda r: r['qb'], reverse=True)
            if fb:
                tier = '兜底'; picks.append(('兜底', fb[0]))
            else:
                # 深兜底: 仅 额≥1亿+开盘<9.8%+非ST+量比∈(5,18)+20涨≤100%, 不要求贴高/趋势
                dp = [r for r in rec['rows'] if r['hard_ok'] and r['amt'] >= AMT_MIN
                      and r['gap'] < GAP_CAP and not r['isST']
                      and (QB_DEEP_LO < r['qb'] < QB_HI) and r['run20'] <= RUN20_CAP_MAIN]
                dp.sort(key=lambda r: r['qb'], reverse=True)
                if dp:
                    tier = '深兜底'; picks.append(('深兜底', dp[0]))

        rec['tier'] = tier
        rec['picks'] = picks
        # 标注每行实现收益(供候选排名), 并记本笔组合收益
        for r in rec['rows']:
            r['ret'] = real_ret(r['c'], T, nxt)
        pick_rets = [real_ret(r['c'], T, nxt) for _, r in picks]
        pick_rets = [x for x in pick_rets if x is not None]
        rec['trade_ret'] = float(np.mean(pick_rets)) if pick_rets else None

    return records, idx_nav, idx_ma20, etf, dview, start_bt, end


# ─────────────────────── 净值与逐年收益 ───────────────────────
def compute_nav(records):
    """两份资金①②错开一天滚动: 每个有买入的进攻日, 由轮值的一份吃下本笔收益。"""
    A, B = 0.5, 0.5
    turn = 0
    series = []     # (T, total_nav, traded?)
    for rec in records:
        if rec['regime'] == 'attack' and rec['picks'] and rec['trade_ret'] is not None:
            r = rec['trade_ret']
            if turn % 2 == 0: A *= (1 + r)
            else: B *= (1 + r)
            turn += 1
            traded = True
        else:
            traded = False
        series.append((rec['T'], A + B, traded))
    return series

def yearly_returns(series, etf):
    nav = pd.Series({t: v for t, v, _ in series})
    years = sorted({t[:4] for t in nav.index})
    rows = []
    prev_end = 1.0
    for y in years:
        ny = nav[[t for t in nav.index if t[:4] == y]]
        if not len(ny):
            continue
        yend = ny.iloc[-1]
        ret = yend / prev_end - 1
        # 参考: 当年 ETF 涨跌
        eref = {}
        for enm, ed in etf.items():
            if ed is None:
                eref[enm] = None; continue
            ey = ed[[t for t in ed.index if t[:4] == y]]
            eref[enm] = (ey['close'].iloc[-1] / ey['close'].iloc[0] - 1) if len(ey) else None
        rows.append(dict(year=y, ret=ret, nav_end=yend,
                         bank=eref.get('银行ETF'), gold=eref.get('黄金ETF')))
        prev_end = yend
    return rows, nav

def max_drawdown(nav):
    peak = nav.cummax()
    return float((nav / peak - 1).min())


# ─────────────────────────── Excel ───────────────────────────
def write_excel(records, series, yrows, nav, etf, dview, start_bt, end, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HDR = PatternFill('solid', fgColor='1F4E78'); HF = Font(color='FFFFFF', bold=True)
    SUB = PatternFill('solid', fgColor='D9E1F2'); SUBF = Font(bold=True, color='1F4E78')
    BUY = PatternFill('solid', fgColor='C6EFCE'); WARN = PatternFill('solid', fgColor='FFC7CE')
    POSF = Font(color='006100'); NEGF = Font(color='9C0006')
    CEN = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='BFBFBF'); BORD = Border(thin, thin, thin, thin)

    wb = Workbook()

    # ── 总览 ──
    ws = wb.active; ws.title = '总览'
    ws['A1'] = 'AI 主池打板回溯 v2.1'; ws['A1'].font = Font(bold=True, size=15, color='1F4E78')
    ws['A2'] = f'区间 {start_bt} ~ {end}  ·  60 只固定池(含幸存者偏差, 账面收益仅供复盘)'
    ws['A2'].font = Font(italic=True, color='808080')
    total = nav.iloc[-1] / 1.0 - 1
    ws['A4'] = '累计收益'; ws['B4'] = total
    ws['A5'] = '最大回撤'; ws['B5'] = max_drawdown(nav)
    traded = [s for s in series if s[2]]
    ws['A6'] = '交易天数'; ws['B6'] = len(traded)
    wins = sum(1 for rec in records if rec.get('trade_ret') is not None and rec['picks'] and rec['trade_ret'] > 0)
    ntr = sum(1 for rec in records if rec.get('trade_ret') is not None and rec['picks'])
    ws['A7'] = '胜率(按笔)'; ws['B7'] = (wins / ntr) if ntr else 0
    for r in range(4, 8):
        ws[f'A{r}'].font = Font(bold=True)
        ws[f'B{r}'].number_format = '0.00%' if r in (4, 5, 7) else '0'

    h = ['年度', '组合收益率', '年末净值', '银行ETF(参考)', '黄金ETF(参考)']
    hr = 9
    for j, c in enumerate(h, 1):
        cell = ws.cell(hr, j, c); cell.fill = HDR; cell.font = HF; cell.alignment = CEN; cell.border = BORD
    for i, yr in enumerate(yrows):
        rr = hr + 1 + i
        vals = [yr['year'], yr['ret'], yr['nav_end'], yr['bank'], yr['gold']]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(rr, j, v); cell.alignment = CEN; cell.border = BORD
            if j in (2, 4, 5) and v is not None:
                cell.number_format = '0.00%'; cell.font = POSF if v >= 0 else NEGF
            if j == 3: cell.number_format = '0.0000'
    ws.column_dimensions['A'].width = 16
    for col in 'BCDE': ws.column_dimensions[col].width = 15

    # 选股规则说明(候选一/二/三 买入条件)
    note_r = hr + len(yrows) + 3
    ws.cell(note_r, 1, '【候选一/二/三 买入条件】').font = Font(bold=True, color='1F4E78')
    notes = [
        '候选基础筛选(均按 T-1 数据): 量比∈(3,18) + 贴高(T-1收盘≥60日最高×93%) + 多头排列(收盘>MA20>MA60) + 成交额(T-1)≥1亿 + 开盘涨幅<9.8% + 非ST, 按量比从大到小排。',
        '候选一: 候选量比第1名 —— 低开照买(20日涨幅≤100% 已由硬过滤保证, 不设开盘涨幅下限)。',
        '候选二: 候选量比第2名, 须 20日涨幅≤80% 且 开盘涨幅>-1.5%; 不满足则该名次空缺, 不顺延不递补。',
        '候选三: 候选量比第3名, 须 20日涨幅≤80% 且 开盘涨幅>-1.5%; 不满足则该名次空缺, 不顺延不递补。',
        '兜底(主选0只): 去量比>3下限与20涨≤80, 留 贴高+多头+额≥1亿+非ST(仍受硬过滤), 取量比第1满仓。',
        '深兜底(主选+兜底0只): 仅 额≥1亿+开盘<9.8%+非ST+量比∈(5,18)+20涨≤100%, 不要求贴高/趋势, 取量比第1满仓; 无量比>5合格票则空仓。',
        '全局硬过滤(任一命中整票剔除): 开盘涨幅>9.8% / 20日涨幅>100% / 量比>20 / 60日最高K为阴线且距今>7日 / 高开不买名单(gap≥0) / 低开不买名单(gap<0)。',
        '高开不买: 英维克、兆易创新、联瑞新材、高澜股份、绿的谐波。  低开不买: 生益电子、风华高科、景旺电子、亨通光电。',
        '执行: 本金分两份错开一天滚动, T开盘买、T+1收盘卖; T收盘跌停的笔改 T+1开盘卖; 防守日(等权指数<MA20)空仓。',
    ]
    for k, t in enumerate(notes):
        ws.cell(note_r + 1 + k, 1, ('· ' if k else '') + t)

    # ── 每年一个工作表 ──
    cols = ['日期', '名称', '量比', '近5日每分钟均量(手)', '20日涨幅', '开盘涨幅', '终值位置',
            '成交额(亿)', '贴高', '多头', '买卖限制', '入选', 'T开盘', 'T+1结算', '收盘获利']
    by_year = {}
    for rec in records:
        by_year.setdefault(rec['T'][:4], []).append(rec)

    for year in sorted(by_year):
        ws = wb.create_sheet(year)
        yr = next((y for y in yrows if y['year'] == year), None)
        ws['A1'] = f'{year} 年 · AI主池打板逐日候选与买入'
        ws['A1'].font = Font(bold=True, size=13, color='1F4E78')
        if yr:
            ws['A2'] = f"组合年度收益 {yr['ret']*100:+.2f}%  ·  年末净值 {yr['nav_end']:.4f}"
            ws['A2'].font = Font(italic=True, bold=True, color='006100' if yr['ret'] >= 0 else '9C0006')
        # 表头
        hr = 4
        for j, c in enumerate(cols, 1):
            cell = ws.cell(hr, j, c); cell.fill = HDR; cell.font = HF; cell.alignment = CEN; cell.border = BORD
        r = hr + 1
        pick_map_tier = {'选1': '候选一', '选2': '候选二', '选3': '候选三',
                         '兜底': '兜底', '深兜底': '深兜底'}
        for rec in by_year[year]:
            T = rec['T']
            if rec['regime'] == 'defend':
                cell = ws.cell(r, 1, f'{T}  防守=空仓(等权指数<MA20, 当日收益0)')
                cell.font = Font(italic=True, color='808080'); r += 1
                continue
            cand_rows = [x for x in rec['rows'] if x['hard_ok'] and x['near'] and x['trend']
                         and x['amt'] >= AMT_MIN and (QB_LO < x['qb'] < QB_HI)
                         and (x['gap'] < GAP_CAP) and not x['isST']]
            if not cand_rows:
                # 展示兜底/深兜底实际买入(若有)
                if rec['picks']:
                    _, pr = rec['picks'][0]
                    cell = ws.cell(r, 1, f"{T}  主选0只 → {rec['tier']}买入: {pr['nm']}(量比{pr['qb']:.1f})")
                else:
                    cell = ws.cell(r, 1, f'{T}  无合格候选 → 空仓')
                cell.font = Font(italic=True, color='808080'); r += 1
                continue
            picks_by_c = {pr['c']: pick_map_tier.get(tag, tag) for tag, pr in rec['picks']}
            cand_rows.sort(key=lambda x: (x['ret'] if x.get('ret') is not None else -9), reverse=True)
            for x in cand_rows:
                limits = []
                if x['run20'] > RUN20_CAP_SUB: limits.append('20涨>80%(只能候选一)')
                if x['c'] in HIGH_OPEN_BAN: limits.append('高开不买')
                if x['c'] in LOW_OPEN_BAN: limits.append('低开不买')
                if x['gap'] <= LOW_OPEN_TOL: limits.append('低开>1.5%(候选二三留空)')
                dv = dview.get(x['c'])
                t_open = dv.loc[T, 'qopen'] if (dv is not None and T in dv.index) else None
                nxt = rec['nxt']
                settle = None
                if dv is not None and nxt and nxt in dv.index:
                    cl_t = dv.loc[T, 'close']; pc_t = dv.loc[T, 'pre_close']
                    pct_t = (cl_t / pc_t - 1) if pc_t else 0
                    lim = -0.198 if is_chuangchuang(x['c']) else -0.098
                    settle = dv.loc[nxt, 'qopen'] if pct_t <= lim else dv.loc[nxt, 'qclose']
                vals = [T, x['nm'], x['qb'], x['permin5'], x['run20'], x['gap'], x['pos'],
                        x['amt'] / 100000, '√' if x['near'] else '', '√' if x['trend'] else '',
                        ' / '.join(limits), picks_by_c.get(x['c'], ''),
                        t_open, settle, x.get('ret')]
                for j, v in enumerate(vals, 1):
                    cell = ws.cell(r, j, v); cell.border = BORD; cell.alignment = CEN
                    if j in (5, 6, 15) and isinstance(v, float): cell.number_format = '0.00%'
                    if j in (3, 4, 7, 8, 13, 14) and isinstance(v, float): cell.number_format = '0.00'
                    if j == 15 and isinstance(v, float):
                        cell.font = POSF if v >= 0 else NEGF
                    if j == 12 and v:
                        cell.fill = BUY; cell.font = Font(bold=True)
                    if j == 11 and v:
                        cell.fill = WARN
                r += 1
            r += 1   # 每个交易日之间空一行
        widths = [11, 10, 7, 18, 10, 10, 9, 10, 6, 6, 26, 9, 9, 9, 10]
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w
        ws.freeze_panes = 'A5'

    wb.save(path)
    print(f'已写出 {path}  ({len(by_year)} 个年度工作表)')


def main():
    records, idx_nav, idx_ma20, etf, dview, start_bt, end = build()
    series = compute_nav(records)
    yrows, nav = yearly_returns(series, etf)
    path = sys.argv[3] if len(sys.argv) > 3 else f'ai_backtest_{start_bt[:4]}_{end[:4]}.xlsx'
    print('\n逐年收益:')
    for y in yrows:
        print(f"  {y['year']}: {y['ret']*100:+7.2f}%   年末净值 {y['nav_end']:.4f}")
    print(f"累计 {nav.iloc[-1]-1:+.2%}  最大回撤 {max_drawdown(nav):+.2%}")
    write_excel(records, series, yrows, nav, etf, dview, start_bt, end, path)

if __name__ == '__main__':
    main()
